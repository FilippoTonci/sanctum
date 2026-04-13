"""Unit tests for the .xlsx reader/writer adapter pair."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sanctum.documents.xlsx_adapter import Reader, Writer

FIXTURE = Path("tests/fixtures/office/invoice.xlsx")


@pytest.fixture()
def reader() -> Reader:
    return Reader()


@pytest.fixture()
def writer() -> Writer:
    return Writer()


def test_reader_only_emits_string_cells(reader: Reader) -> None:
    doc = reader.read(FIXTURE)
    assert doc.format == "xlsx"
    assert doc.segments, "expected string cells in fixture"
    # Numerics must not have become segments
    assert all(seg.id.startswith("sheet=") for seg in doc.segments)
    # Fixture has 'INVOICE' as A1 on the Header sheet
    assert any(seg.text == "INVOICE" for seg in doc.segments)


def test_reader_metadata_records_sheet(reader: Reader) -> None:
    doc = reader.read(FIXTURE)
    seg = next(s for s in doc.segments if s.text == "INVOICE")
    assert seg.metadata["sheet"] == "Header"
    assert seg.id == "sheet=Header/A1"


def test_round_trip_preserves_numeric_cells(
    reader: Reader, writer: Writer, tmp_path: Path
) -> None:
    doc = reader.read(FIXTURE)
    out = tmp_path / "roundtrip.xlsx"
    writer.write(doc, out)

    original = load_workbook(str(FIXTURE))
    rewritten = load_workbook(str(out))
    for sheet_name in original.sheetnames:
        for row_a, row_b in zip(
            original[sheet_name].iter_rows(values_only=True),
            rewritten[sheet_name].iter_rows(values_only=True),
            strict=True,
        ):
            assert row_a == row_b


def test_writer_applies_edits(reader: Reader, writer: Writer, tmp_path: Path) -> None:
    doc = reader.read(FIXTURE)
    target = next(seg for seg in doc.segments if seg.text == "INVOICE")
    new_segments = [
        seg.model_copy(update={"text": "<REDACTED>"}) if seg.id == target.id else seg
        for seg in doc.segments
    ]
    mutated = doc.model_copy(update={"segments": new_segments})
    mutated.raw_handle = doc.raw_handle

    out = tmp_path / "edited.xlsx"
    writer.write(mutated, out)

    wb = load_workbook(str(out))
    assert wb["Header"]["A1"].value == "<REDACTED>"


def test_writer_rejects_missing_raw_handle(writer: Writer, reader: Reader, tmp_path: Path) -> None:
    doc = reader.read(FIXTURE)
    doc.raw_handle = None
    with pytest.raises(ValueError, match="raw_handle"):
        writer.write(doc, tmp_path / "x.xlsx")


def test_registry_dispatches_to_xlsx_adapter() -> None:
    from sanctum.documents import adapter_for

    r, w = adapter_for(FIXTURE)
    assert isinstance(r, Reader)
    assert isinstance(w, Writer)


def test_skips_missing_sheet_silently(writer: Writer, reader: Reader, tmp_path: Path) -> None:
    """If a segment refers to a sheet the workbook no longer has, skip it."""
    src = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "keep"
    wb.save(str(src))

    doc = reader.read(src)
    doc.segments.append(
        type(doc.segments[0])(id="sheet=Ghost/A1", text="phantom", metadata={})
    )
    out = tmp_path / "out.xlsx"
    writer.write(doc, out)
    assert load_workbook(str(out))["Data"]["A1"].value == "keep"


def test_invalid_segment_id_raises(writer: Writer) -> None:
    with pytest.raises(ValueError, match="Invalid xlsx segment id"):
        Writer._parse_segment_id("bogus")
