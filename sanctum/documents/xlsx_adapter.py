"""Excel (.xlsx) structured reader/writer backed by openpyxl.

Only string cells are emitted as segments — numeric, boolean, date, and
formula cells cannot carry free-text PII and stringifying them would
corrupt display formatting on round-trip. The raw handle is an
openpyxl ``Workbook`` so the writer preserves styles, merges, and any
other feature we don't currently project through segments.

Segment IDs use ``sheet={name}/{coord}`` (e.g. ``sheet=Header/A3``).
Excel disallows ``/`` in sheet names, so the delimiter is unambiguous.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from sanctum.documents.structured import build_document, build_segment

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook

    from sanctum.core.models import StructuredDocument


_STRING_DATA_TYPE = "s"


class Reader:
    """Emit one segment per string-typed cell, in sheet-then-row-then-col order."""

    def read(self, path: Path) -> StructuredDocument:
        workbook: Workbook = load_workbook(str(path))
        segments = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type != _STRING_DATA_TYPE:
                        continue
                    if cell.value is None:
                        continue
                    seg_id = f"sheet={sheet_name}/{cell.coordinate}"
                    segments.append(build_segment(seg_id, str(cell.value), sheet=sheet_name))

        return build_document(
            source_path=path,
            fmt="xlsx",
            segments=segments,
            raw_handle=workbook,
        )


class Writer:
    """Project edits back onto the raw :class:`Workbook` and save."""

    def write(self, doc: StructuredDocument, path: Path) -> None:
        workbook: Workbook | None = doc.raw_handle
        if workbook is None:
            raise ValueError(
                "XlsxWriter requires the StructuredDocument.raw_handle from XlsxReader"
            )

        for segment in doc.segments:
            sheet_name, coord = self._parse_segment_id(segment.id)
            if sheet_name not in workbook.sheetnames:
                continue
            workbook[sheet_name][coord] = segment.text

        workbook.save(str(path))

    @staticmethod
    def _parse_segment_id(seg_id: str) -> tuple[str, str]:
        prefix, _, coord = seg_id.partition("/")
        if not prefix.startswith("sheet=") or not coord:
            raise ValueError(f"Invalid xlsx segment id: {seg_id!r}")
        return prefix.removeprefix("sheet="), coord
